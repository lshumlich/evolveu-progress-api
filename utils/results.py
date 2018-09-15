"""

pipenv shell
python3 -m utils.results
DATABASE_URL="dbname=larry2 user=larry" python3 -m utils.results

"""
import json
import datetime

import sql.sql
import openpyxl

def get_results():

	the_day = '2018-09-10'
	_s = 'Run Date: ' + str(datetime.date.today())


	wb = openpyxl.Workbook()
	ws = wb.create_sheet(the_day, 0)

	row = 1
	col = 1

	ws.cell(row=row, column=col).value = 'Date'
	col += 1

	ws.cell(row=row, column=col).value = 'Name'
	col += 1

	questions = sql.sql.get_questions()
	for q in questions:
		ws.cell(row=row, column=col).value = q['code']
		col += 1

	ws.cell(row=row, column=col).value = 'Going Well'
	col += 1

	ws.cell(row=row, column=col).value = 'Issues'
	col += 1

	ws.cell(row=row, column=col).value = 'What to Try'
	col += 1

	results = sql.sql.get_results()
	for r in results:
		row += 1
		col = 1
		ws.cell(row=row, column=col).value = r[0]
		col += 1
		ws.cell(row=row, column=col).value = r[1]
		re = json.loads(r[2])
		# re = r[1]
		for q in questions:
			code = q['code']
			col += 1
			if code in re:
				value = re[code]
			else:
				value = 0

			ws.cell(row=row, column=col).value = value

		col += 1
		ws.cell(row=row, column=col).value = r[3]
		col += 1
		ws.cell(row=row, column=col).value = r[4]
		col += 1
		ws.cell(row=row, column=col).value = r[5]

	return wb

def save_results():
	wb = get_results()
	wb.save('results.xlsx')

def test():
	None



save_results()